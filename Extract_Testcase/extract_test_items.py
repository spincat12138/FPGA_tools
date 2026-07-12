#!/usr/bin/env python3
"""Extract summarized test items from FPGA-related text logs."""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, Iterator, List, Optional


ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "latin-1")

TABLE_SUITE_RE = re.compile(r"^\s*<([^<>]+)>\s*$")
TESTSUITE_RE = re.compile(r"^=+\s+Started Testsuite\s+(.+?)\s+=+\s*$")
TEST_NAME_RE = re.compile(r"Test Name:\s*(.+?):\s*(PASSED|FAILED|SKIPPED|ERROR)", re.IGNORECASE)
STATUS_TOKENS = {"PASS", "FAIL", "PASSED", "FAILED", "SKIPPED", "ERROR"}
PLATFORMS = ("ultra", "93k", "j750")
SUMMARY_FIELDS = ("test_item", "occurrences", "first_line", "last_line")


@dataclass(frozen=True)
class ParsedTestItem:
    name: str
    line_no: int


@dataclass
class TestItemSummary:
    test_item: str
    occurrences: int
    first_line: int
    last_line: int


def read_text(path: Path) -> str:
    last_error = None  # type: Optional[Exception]
    for encoding in ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    return path.read_text()


def is_number(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def is_ultra_result_row(tokens: List[str], has_suite_context: bool) -> bool:
    return (
        has_suite_context
        and len(tokens) >= 3
        and tokens[0].isdigit()
        and tokens[1].isdigit()
        and not is_number(tokens[2])
    )


def is_j750_result_row(tokens: List[str]) -> bool:
    return (
        len(tokens) >= 5
        and tokens[0].isdigit()
        and tokens[1].isdigit()
        and tokens[2].upper() in STATUS_TOKENS
    )


def extract_ultra_items(lines: List[str]) -> Iterator[ParsedTestItem]:
    has_suite_context = False
    for line_no, line in enumerate(lines, start=1):
        if TABLE_SUITE_RE.match(line):
            has_suite_context = True
            continue

        tokens = line.split()
        if is_ultra_result_row(tokens, has_suite_context):
            yield ParsedTestItem(name=tokens[2], line_no=line_no)


def extract_j750_items(lines: List[str]) -> Iterator[ParsedTestItem]:
    for line_no, line in enumerate(lines, start=1):
        tokens = line.split()
        if is_j750_result_row(tokens):
            yield ParsedTestItem(name=tokens[3], line_no=line_no)


def normalize_93k_test_name(name: str) -> str:
    normalized = re.sub(r"\[\d+\]$", "", name.strip())
    return normalized.split("@", 1)[0]


def extract_93k_items(lines: List[str]) -> Iterator[ParsedTestItem]:
    for line_no, line in enumerate(lines, start=1):
        stripped = line.strip()
        suite_match = TESTSUITE_RE.match(stripped)
        if suite_match:
            yield ParsedTestItem(name=suite_match.group(1).strip(), line_no=line_no)
            continue

        test_match = TEST_NAME_RE.search(stripped)
        if test_match:
            yield ParsedTestItem(name=normalize_93k_test_name(test_match.group(1)), line_no=line_no)


PARSER_BY_PLATFORM = {
    "ultra": extract_ultra_items,
    "93k": extract_93k_items,
    "j750": extract_j750_items,
}  # type: Dict[str, Callable[[List[str]], Iterator[ParsedTestItem]]]


def extract_file(path: Path, platform: str) -> List[ParsedTestItem]:
    text = read_text(path)
    lines = text.splitlines()
    parser = PARSER_BY_PLATFORM.get(platform)
    if parser is None:
        raise ValueError(f"不支持的平台格式: {platform}")
    return list(parser(lines))


def summarize_items(items: Iterable[ParsedTestItem]) -> List[TestItemSummary]:
    groups = {}  # type: Dict[str, TestItemSummary]
    for item in items:
        test_item = item.name.strip()
        if not test_item:
            continue

        summary = groups.get(test_item)
        if summary is None:
            groups[test_item] = TestItemSummary(
                test_item=test_item,
                occurrences=1,
                first_line=item.line_no,
                last_line=item.line_no,
            )
            continue

        summary.occurrences += 1
        summary.first_line = min(summary.first_line, item.line_no)
        summary.last_line = max(summary.last_line, item.line_no)

    return sorted(groups.values(), key=lambda row: (row.first_line, row.test_item))


def iter_input_files(inputs: List[Path]) -> List[Path]:
    files = []  # type: List[Path]
    for input_path in inputs:
        if input_path.is_dir():
            files.extend(sorted(input_path.glob("*.txt")))
        elif input_path.is_file():
            files.append(input_path)
        else:
            raise FileNotFoundError(f"路径不存在: {input_path}")
    return files


def write_excel(path: Path, items: List[TestItemSummary]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel 输出需要安装 openpyxl：python -m pip install openpyxl") from exc

    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "test_items"
    sheet.append(SUMMARY_FIELDS)
    for item in items:
        sheet.append([getattr(item, field) for field in SUMMARY_FIELDS])

    header_fill = PatternFill(fill_type="solid", fgColor="92D050")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="提取 data 文本文件中的测试项")
    parser.add_argument(
        "inputs",
        nargs="*",
        type=Path,
        default=[Path("data")],
        help="输入 txt 文件或目录，默认 data",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出目录；单文件输入时也可指定具体输出文件，默认输出到输入文件所在目录",
    )
    parser.add_argument(
        "-p",
        "--platform",
        required=True,
        choices=PLATFORMS,
        help="显式指定输入文件所属平台格式",
    )
    return parser.parse_args()


def output_path_for(input_file: Path, output: Optional[Path], multiple: bool) -> Path:
    suffix = ".xlsx"
    if output is None:
        return input_file.parent / f"{input_file.stem}_test_items{suffix}"
    if output.suffix and not multiple:
        return output.with_suffix(suffix)
    output_dir = output if not output.suffix else output.parent
    return output_dir / f"{input_file.stem}_test_items{suffix}"


def main() -> int:
    args = parse_args()
    files = iter_input_files(args.inputs)
    if not files:
        print("未找到 txt 文件")
        return 1

    for file in files:
        items = extract_file(file, args.platform)
        output_items = summarize_items(items)
        output_path = output_path_for(file, args.output, multiple=len(files) > 1)
        write_excel(output_path, output_items)
        print(f"{file}: platform={args.platform}, items={len(items)}, rows={len(output_items)} -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
